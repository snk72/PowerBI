import openpyxl
import os

# set the path to the directory where the xls file is located
xls_dir_path = r"C:\Users\sindre\OneDrive - BRAbank ASA\Documents - BRAbank Likviditet\General\Collateral\LEA_BANKFX.xls"


# set the path to the xlsx file you want to create
xlsx_file_path = r"C:\Users\sindre\OneDrive - BRAbank ASA\Documents - BRAbank Likviditet\General\Collateral\XLSX\LEA_BANKFX.xlsx"

# check if the xlsx file already exists and delete it if it does
if os.path.exists(xlsx_file_path):
    os.remove(xlsx_file_path)

# load the xls file and create a new xlsx file
workbook = openpyxl.load_workbook(xls_dir_path)
new_workbook = openpyxl.Workbook()

# loop through each sheet in the xls workbook and copy it to the new xlsx workbook
for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    new_worksheet = new_workbook.create_sheet(sheet_name)

    # copy each cell from the xls worksheet to the xlsx worksheet
    for row in worksheet.iter_rows():
        for cell in row:
            new_worksheet[cell.coordinate].value = cell.value

# save the new xlsx workbook to disk
new_workbook.save(xlsx_file_path)